import os
import sys
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.styles import Font, Alignment
import subprocess
import platform

# Header ASCII
HEADER = r"""
 _   _ ___  ___  ___  ______            _______   _______  _____ _     
| \ | ||  \/  | / _ \ | ___ \          |  ___\ \ / /  __ \|  ___| |    
|  \| || .  . |/ /_\ \| |_/ /  ______  | |__  \ V /| /  \/| |__ | |    
| . ` || |\/| ||  _  ||  __/  |______| |  __| /   \| |    |  __|| |    
| |\  || |  | || | | || |              | |___/ /^\ \ \__/\| |___| |____
\_| \_/\_|  |_/\_| |_/\_|              \____/\/   \/\____/\____/\_____/
"""

def print_help():
    help_text = """
Usage: python3 script.py <IP_or_Domain> <output.xlsx>

This script performs a nmap scan on the specified IP or domain, and saves the results in an Excel file.
    
Arguments:
    <IP_or_Domain>    IP address, CIDR block, or domain to scan
    <output.xlsx>     Path to the output Excel file
"""
    print(HEADER)
    print(help_text)

def run_nmap(ip_or_domain):
    output_dir = ip_or_domain.split('/')[0]
    output_file_prefix = ip_or_domain.split('/')[1] if '/' in ip_or_domain else ip_or_domain
    output_file_base = os.path.join(output_dir, output_file_prefix)
    os.makedirs(output_dir, exist_ok=True)

    command = f"nmap -Pn -A -oA {output_file_base} {ip_or_domain}"

    # Usar subprocess para ejecutar el comando de manera compatible
    try:
        result = subprocess.run(command, shell=True, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        print(result.stdout.decode())
        print(result.stderr.decode())
    except subprocess.CalledProcessError as e:
        print(f"Error ejecutando nmap: {e}")
        sys.exit(1)
    
    return f"{output_file_base}.xml"

def parse_nmap_file(nmap_file):
    tree = ET.parse(nmap_file)
    root = tree.getroot()
    scan_results = []

    for host in root.findall('host'):
        ip_addr = host.find('address').get('addr') if host.find('address') is not None else 'N/A'
        hostname = 'N/A'
        for hostname_elem in host.findall(".//elem[@key='NetBIOS_Computer_Name']"):
            hostname = hostname_elem.text
        os_guess = ', '.join([f"{osmatch.get('name')} ({osmatch.get('accuracy')}%)" for osmatch in host.findall(".//osmatch")[:2]])

        ports_info = []
        for port in host.findall(".//port"):
            port_num = port.get('portid')
            state = port.find('state').get('state')
            service_elem = port.find('service')
            service = service_elem.get('name') if service_elem is not None else 'N/A'
            version_info = (service_elem.get('product', '') + ' ' + service_elem.get('version', '')).strip() if service_elem is not None else ''
            ports_info.append((port_num, state, service, version_info))

        scan_results.append({
            'hostname': hostname,
            'ip': ip_addr,
            'ports': ports_info,
            'os': os_guess
        })
    return scan_results

def create_excel(scan_results, output_file):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Nmap Scan Results"
    ws.append(["Hostname", "IP", "PUERTOS", "ESTADOS", "SERVICIOS", "VERSIONES", "S.O"])

    for result in scan_results:
        hostname = result['hostname']
        ip_addr = result['ip']
        ports = result['ports']
        os_guess = result['os']

        if not ports:
            ws.append([hostname, ip_addr, "", "", "", "", os_guess])
            continue

        first_row = ws.max_row + 1
        for port, state, service, version in ports:
            ws.append(["", "", port, state, service, version, ""])

        last_row = ws.max_row
        ws.cell(row=first_row, column=1, value=hostname)
        ws.cell(row=first_row, column=2, value=ip_addr)
        ws.cell(row=first_row, column=7, value=os_guess)

        ws.merge_cells(start_row=first_row, start_column=1, end_row=last_row, end_column=1)
        ws.merge_cells(start_row=first_row, start_column=2, end_row=last_row, end_column=2)
        ws.merge_cells(start_row=first_row, start_column=7, end_row=last_row, end_column=7)

        # Alineación al centro vertical y horizontal
        for row in ws.iter_rows(min_row=first_row, max_row=last_row, min_col=1, max_col=7):
            for cell in row:
                cell.alignment = Alignment(vertical='center', horizontal='center')

    wb.save(output_file)

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print_help()
        sys.exit(1)

    print(HEADER)
    
    ip_or_domain = sys.argv[1]
    output_file = sys.argv[2]

    xml_file = run_nmap(ip_or_domain)
    scan_results = parse_nmap_file(xml_file)
    create_excel(scan_results, output_file)
